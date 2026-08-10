#!/usr/bin/env python3
"""ETL for the PSOBM mental-health indicators dashboard.

This pipeline is intentionally decoupled from index.html. It produces a candidate
JSON artifact and never overwrites the current demonstrative frontend by default.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import sys
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

REQUIRED_COLUMNS = {
    "period",
    "monitoring_records",
    "assessments_completed",
    "cycle_days_total",
    "cycle_count",
    "attendance_records",
    "unique_beneficiaries",
    "referrals_completed",
    "referrals_total",
    "urgent_demands",
    "effective_strength",
    "response_hours_total",
    "response_count",
    "mental_health_leave_count",
    "lost_days",
}
SENSITIVE_TOKENS = {
    "name", "nome", "cpf", "rg", "email", "e-mail", "matricula", "matrícula",
    "phone", "telefone", "address", "endereco", "endereço", "patient", "paciente",
}

@dataclass
class Indicator:
    id: str
    group: str
    name: str
    value: float
    unit: str
    period: str


def n(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        number = float(str(value).replace(",", "."))
    except ValueError as exc:
        raise ValueError(f"non-numeric value: {value!r}") from exc
    if not math.isfinite(number):
        raise ValueError(f"non-finite value: {value!r}")
    return number


def safe_div(num: float, den: float, factor: float = 1.0) -> float:
    return 0.0 if den == 0 else num / den * factor


def validate_headers(headers: Iterable[str]) -> None:
    normalized = {h.strip() for h in headers if h}
    sensitive = sorted(h for h in normalized if h.casefold() in SENSITIVE_TOKENS)
    if sensitive:
        raise ValueError("identifiable/sensitive columns are not allowed in this aggregated pipeline: " + ", ".join(sensitive))
    missing = sorted(REQUIRED_COLUMNS - normalized)
    if missing:
        raise ValueError("missing required aggregate columns: " + ", ".join(missing))


def calculate(row: dict[str, Any]) -> list[Indicator]:
    period = str(row["period"]).strip()
    strength = n(row["effective_strength"])
    return [
        Indicator("i1", "cronico", "Taxa de conclusão do monitoramento PSOBM", round(safe_div(n(row["assessments_completed"]), n(row["monitoring_records"]), 100), 1), "%", period),
        Indicator("i2", "cronico", "Tempo médio do ciclo de monitoramento", round(safe_div(n(row["cycle_days_total"]), n(row["cycle_count"])), 1), "dias", period),
        Indicator("i3", "cronico", "Índice de recorrência assistencial por beneficiário", round(safe_div(n(row["attendance_records"]), n(row["unique_beneficiaries"])), 2), "razão", period),
        Indicator("i4", "cronico", "Taxa de continuidade assistencial concluída", round(safe_div(n(row["referrals_completed"]), n(row["referrals_total"]), 100), 1), "%", period),
        Indicator("i5", "agudo", "Taxa de demandas psicológicas urgentes por 1.000 militares", round(safe_div(n(row["urgent_demands"]), strength, 1000), 1), "por 1.000", period),
        Indicator("i6", "agudo", "Tempo médio até acolhimento", round(safe_div(n(row["response_hours_total"]), n(row["response_count"])), 1), "horas", period),
        Indicator("i7", "agudo", "Taxa de afastamento psicológico/psiquiátrico", round(safe_div(n(row["mental_health_leave_count"]), strength, 100), 1), "%", period),
        Indicator("i8", "agudo", "Dias perdidos por afastamento psicológico/psiquiátrico", round(n(row["lost_days"]), 1), "dias", period),
    ]


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        validate_headers(reader.fieldnames or [])
        return list(reader)


def read_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list) or not data:
        raise ValueError("JSON source must be a non-empty array of aggregate rows")
    validate_headers(data[0].keys())
    return data


def read_xlsx(path: Path, sheet_name: str = "") -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit("openpyxl is required for XLSX sources") from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    iterator = ws.iter_rows(values_only=True)
    headers = [str(v or "").strip() for v in next(iterator)]
    validate_headers(headers)
    rows = []
    for values in iterator:
        row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        if str(row.get("period", "")).strip():
            rows.append(row)
    return rows


def load(path: Path, sheet: str = "") -> list[dict[str, Any]]:
    suffix = path.suffix.casefold()
    if suffix == ".csv": return read_csv(path)
    if suffix == ".json": return read_json(path)
    if suffix in {".xlsx", ".xlsm"}: return read_xlsx(path, sheet)
    raise ValueError("supported source formats: .csv, .json, .xlsx, .xlsm")


def build(args: argparse.Namespace) -> int:
    rows = load(Path(args.source), args.sheet)
    series: dict[str, list[dict[str, Any]]] = {f"i{i}": [] for i in range(1, 9)}
    for row in rows:
        for indicator in calculate(row):
            series[indicator.id].append(asdict(indicator))

    payload = {
        "schemaVersion": 1,
        "privacyMode": "aggregated-non-identifiable",
        "sourceRows": len(rows),
        "indicators": series,
    }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"candidate dashboard data generated at {out}; frontend remains unchanged")
    return 0


def self_test(_: argparse.Namespace) -> int:
    row = {
        "period":"2026-01", "monitoring_records":100, "assessments_completed":75,
        "cycle_days_total":180, "cycle_count":10, "attendance_records":142,
        "unique_beneficiaries":100, "referrals_completed":68, "referrals_total":100,
        "urgent_demands":127, "effective_strength":10000,
        "response_hours_total":432, "response_count":10,
        "mental_health_leave_count":210, "lost_days":1235,
    }
    values = {i.id: i.value for i in calculate(row)}
    expected = {"i1":75.0,"i2":18.0,"i3":1.42,"i4":68.0,"i5":12.7,"i6":43.2,"i7":2.1,"i8":1235.0}
    if values != expected:
        print(json.dumps({"expected": expected, "actual": values}, indent=2), file=sys.stderr)
        return 1
    print("self-test passed: all eight formulas are stable")
    return 0


def parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(description="Aggregated ETL for PSOBM dashboard")
    sub = p.add_subparsers(dest="command", required=True)
    s = sub.add_parser("self-test")
    s.set_defaults(func=self_test)
    b = sub.add_parser("build")
    b.add_argument("--source", required=True)
    b.add_argument("--sheet", default="")
    b.add_argument("--output", default="artifacts/dashboard.generated.json")
    b.set_defaults(func=build)
    return p


def main() -> int:
    args = parser().parse_args()
    return int(args.func(args))

if __name__ == "__main__":
    sys.exit(main())
